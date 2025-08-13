from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect
from .forms import ExcelUploadForm
from .models import ExcelUpload, DTRRecord, ProjectListRecord, ProjectManpowerRecord, Sheet2Record
import openpyxl
from datetime import datetime
from django.contrib import messages

@login_required
def client_dashboard(request):
    upload_status = ""

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            upload = form.save(commit=False)
            upload.client = request.user
            upload.uploaded_by = request.user
            upload.original_filename = request.FILES['file'].name
            upload.save()

            wb = openpyxl.load_workbook(request.FILES['file'], data_only=True)

            for index, sheet in enumerate(wb.worksheets):
                sheet_name = sheet.title
                max_row = sheet.max_row
                max_col = sheet.max_column

                if index == 0:
                    # First sheet – DTR data with headers
                    for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True):
                        if not row[0]:
                            continue
                        DTRRecord.objects.create(
                            upload=upload,
                            sheet_name=sheet_name,
                            employee_code=row[0],
                            employee_name=row[1],
                            total_hours=row[2],
                            nd_reg_hours=row[3],
                            absences=row[4],
                            tardiness=row[5],
                            undertime=row[6],
                            ot_regular=row[7],
                            nd_ot_reg=row[8],
                            ot_restday=row[9],
                            nd_restday=row[10],
                            ot_rest_excess=row[11],
                            nd_rest_excess=row[12],
                            ot_special_holiday=row[13],
                            nd_special_holiday=row[14],
                            ot_sh_excess=row[15],
                            nd_sh_excess=row[16],
                            ot_legal_holiday=row[17],
                            special_holiday=row[18],
                            ot_legal_excess=row[19],
                            nd_legal_excess=row[20],
                            ot_sh_rest=row[21],
                            nd_sh_rest=row[22],
                            ot_sh_rest_excess=row[23],
                            nd_sh_rest_excess=row[24],
                            legal_rest=row[25],
                            nd_legal_rest=row[26],
                            ot_legal_rest_excess=row[27],
                            nd_legal_rest_excess=row[28],
                            vac_leave_applied=row[29],
                            sick_leave_applied=row[30],
                            backpay_vl=row[31],
                            backpay_sl=row[32],
                            ot_regular_excess=row[33],
                            nd_ot_reg_excess=row[34],
                            legal_holiday=row[35],
                            nd_legal_holiday=row[36],
                            overnight_rate=row[37],
                            project=row[38],
                        )
                elif index == 1:
                    # Sheet 2 – No header
                    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                        Sheet2Record.objects.create(
                            upload=upload,
                            sheet_name=sheet_name,
                            col_a=row[0],
                            col_b=row[1],
                            col_c=row[2],
                            # Add more if needed
                        )
                elif index == 2:
                    # Sheet 3 – With fixed headers like "PROJECT", "SUPERVISOR", etc.
                    for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True):
                        if not row[0]:  # Skip rows without a 'project' value
                            continue
                        ProjectManpowerRecord.objects.create(
                            upload=upload,
                            sheet_name=sheet_name,
                            project=row[0],
                            supervisor=row[1],
                            tl=row[2],
                            messenger=row[3],
                            hk=row[4],
                            gnl=row[5],
                            day_off_reliever=row[6],
                            aa=row[7],
                            electrician=row[8],
                            plumber=row[9],
                            driver=row[10],
                            total=row[11],
                        )
                elif index == 3:
                    # Sheet 4 – Only 'project'
                    for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True):
                        ProjectListRecord.objects.create(
                            upload=upload,
                            sheet_name=sheet_name,
                            project=row[0]
                        )

            upload_status = "✅ Upload and data import successful!"
            form = ExcelUploadForm()  # Reset form
        else:
            upload_status = "❌ Upload failed: Invalid file."
    else:
        form = ExcelUploadForm()

    uploads = ExcelUpload.objects.filter(client=request.user).order_by('-uploaded_at')

    return render(request, 'core/client_dashboard.html', {
        'form': form,
        'upload_status': upload_status,
        'uploads': uploads,
    })


# Helper: only allow admin users
def is_admin(user):
    return user.is_superuser  # or use your custom condition

@user_passes_test(is_admin)
def admin_dashboard(request):
    if not request.user.is_superuser:
        return redirect('client_dashboard')

    records = DTRRecord.objects.select_related('upload').order_by('id')
    sheet2_records = Sheet2Record.objects.select_related('upload').order_by('id')
    sheet3_records = ProjectManpowerRecord.objects.select_related('upload').order_by('id')
    sheet4_records = ProjectListRecord.objects.select_related('upload').order_by('id')

    uploaded_files = ExcelUpload.objects.all().order_by('-uploaded_at')

    # Collect distinct sheet names from all record types
    dtr_sheets = DTRRecord.objects.values_list('sheet_name', flat=True)
    sheet2_sheets = Sheet2Record.objects.values_list('sheet_name', flat=True)
    sheet3_sheets = ProjectManpowerRecord.objects.values_list('sheet_name', flat=True)
    sheet4_sheets = ProjectListRecord.objects.values_list('sheet_name', flat=True)

    # Combine and deduplicate
    all_sheet_names = list(set(dtr_sheets.union(sheet2_sheets, sheet3_sheets, sheet4_sheets)))

    return render(request, 'core/admin_dashboard.html', {
        'records': records,
        'sheet2_records': sheet2_records,
        'sheet3_records': sheet3_records,
        'sheet4_records': sheet4_records,
        'uploaded_files': uploaded_files,
        'sheet_names': all_sheet_names,
    })

@login_required
def clear_data(request):
    if request.method == 'POST':
        DTRRecord.objects.all().delete()
        ProjectManpowerRecord.objects.all().delete()
        ProjectListRecord.objects.all().delete()
        Sheet2Record.objects.all().delete()
        messages.success(request, "All records have been deleted.")
    return redirect('admin_dashboard')